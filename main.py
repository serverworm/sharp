import logging
import openpyxl
from datetime import datetime
from collecting_the_urls_card_of_the_inventory_of_the_search_references import start_collecting
from collecting_information_from_the_card import collected_info

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)


def create_excel(data, filename=f'logs/{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'):
    # Создаем новую книгу и лист
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Задаем заголовки столбцов
    headers = ["Название карточки", "Ссылка на карточку", "Цена"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Заполняем данными
    for row_num, row_data in enumerate(data, start=2):
        print(f"Processing row {row_num}: {row_data}")
        if row_data is None:
            print(f"Warning: Row {row_num} is None and will be skipped.")
            continue
        for col_num, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)
        print(f"Row {row_num} written with data: {row_data}")

    # Сохраняем файл
    workbook.save(filename)
    print(f"File '{filename}' has been saved successfully.")

    # Сохраняем файл
    workbook.save(filename)
    print(f"File '{filename}' has been saved successfully.")


def main(search_url):
    urls_cards = start_collecting(search_url)
    logging.info('Сбор ссылок на карточки из поисковой выдачи успешно закончен.')
    result = []
    for collected_url in urls_cards:
        result.append(collected_info(collected_url))
    create_excel(result)


if __name__ == '__main__':
    # тестовый url на поисковую выдачу по запросу "Балтика №0 Нефильтрованное Пшеничное Банка 0,45" по адресу Гастелло 2, Москва
    url = 'https://www.ozon.ru/search/?text=Flash+Up+Energy+%D0%B1%D0%B0%D0%BD%D0%BA%D0%B0+0%2C45&from_global=true&__rr=1'
    main(url)
